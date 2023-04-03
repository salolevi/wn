import fmeobjects
import fme
from openpyxl import load_workbook
from datetime import datetime
import time
import os
from scipy.spatial import KDTree
from scipy.spatial.distance import cdist
from shapely.geometry import Point, LineString

SITIO = ["SITIOS"]
CAMARA = ["CAMARAS"]
RED = ["RED"]
TRAMO = ["TRAMO TRONCAL", "TRAMO PROVINCIAL", "TRAMO SILICA"]
SUBTRAMO = ["SUBTRAMO TRONCAL", "SUBTRAMO PROVINCIAL"]
GABINETE = ["GABINETE","GABINETES"]
SHELTER = ["SHELTER"]
CASETA = ["CASETA"]
EDT = ["EDT´S"]
FRONTERA_OPTICA = ["FRONTERAS OPTICAS", "FRONTERA ÓPTICA"]
POSTE = ["POSTES"]
TRITUBO = ["TRITUBO"]
ELEMENT_BASENAMES = ["Placemark", "Perdriel_postes01", "postes_agrelo_1", "postes_elmanzano_02",
             "Shelters_Transporte", "FO_Existente_GSJ"]
CARPETAS_PRINCIPALES = SITIO + CAMARA + RED + FRONTERA_OPTICA + POSTE
TIPOS_SITIO = GABINETE + SHELTER + CASETA + EDT

PROVINCIA = "CATAMARCA"
SITIOS_EXCEL_PATH = r"C:/Users/j2sae/OneDrive/Escritorio/Trabajo/B3/Migracion/SITIOS - (Catamarca) -migracion.xlsx"
ENTORNO_EXCEL_PATH = r"C:/Users/j2sae/OneDrive/Escritorio/Trabajo/B3/Migracion/Entorno Bloque 3.xlsx"
RTU_EXCEL_PATH = r"C:/Users/j2sae/OneDrive/Escritorio/Trabajo/B3/Migracion/RTU de Bloque 3.xlsx"
TRAMOS_Y_SUBTRAMOS_EXCEL_PATH = r"C:/Users/j2sae/OneDrive/Escritorio/Trabajo/B3/Migracion/ATRIBUTOS TRAMOS Y DERIVACIONES -migracion.xlsx"
RUTA_MIGRACIONES = r"C:/Users/j2sae/OneDrive/Escritorio/Trabajo/B3/Migracion"

OBJECT_ID_INICIAL = 1

BUFFERS_PER_TYPE = {
    "0": 1,
    "1": 2,
    "2": 4,
    "3": 6,
    "4": 8,
    "5": 12,
    "6": 24
}

FIBER_PER_TYPE = {
    "8": 8,
    "12": 12,
    "36": 36,
    "48": 48,
    "72": 72,
    "96": 96,
    "144": 144,
    "288": 288
}

buffer_colors = ["AZ","NR","VR","MR","GR","BL","RJ","NE","AM","VL","RS","CE",
                 "AZ-R","NR-R","VR-R","MR-R","GR-R","BL-R","RJ-R","NE-R","AM-R","VL-R","RS-R","CE-R"]

fiber_colors = ["AZ","NR","VR","MR","GR","BL","RJ","NE","AM","VL","RS","CE"]

INFRA_NETWORK_IDS = {
    "TUCUMAN": "32",
    "SALTA": "30",
    "JUJUY": "28",
    "CATAMARCA": "26",
    "LA RIOJA": "24",
    "NEUQUEN": "22",
    "SAN JUAN": "20",
    "MENDOZA": "18",
    "RIO NEGRO": "16",
    "USHUAIA": "14",
    "SANTA CRUZ": "12",
    "CHUBUT": "10"
}

FO_NETWORK_IDS = {
    "TUCUMAN": "33",
    "SALTA": "31",
    "JUJUY": "29",
    "CATAMARCA": "27",
    "LA RIOJA": "25",
    "NEUQUEN": "23",
    "SAN JUAN": "21",
    "MENDOZA": "19",
    "RIO NEGRO": "17",
    "USHUAIA": "15",
    "SANTA CRUZ": "13",
    "CHUBUT": "11"
}

PROVINCIAS_ENTORNO = {
    "Tucu": "TUCUMAN",
    "Jujuy": "JUJUY",
    "Salta": "SALTA",
    "Cata": "CATAMARCA",
    "LRioja": "LA RIOJA",
    "Neuq": "NEUQUEN",
    "SJuan": "SAN JUAN",
    "Mend": "MENDOZA",
    "RNegro": "RIO NEGRO",
    "TFuego": "USHUAIA",
    "SCruz": "SANTA CRUZ",
    "Chub": "CHUBUT"
}

# dependiendo de la provincia, devuelve el network_id
def obtener_network_id_infra():
    return INFRA_NETWORK_IDS[PROVINCIA]

# dependiendo de la provincia, devuelve el network_id
def obtener_network_id_fo():
    return FO_NETWORK_IDS[PROVINCIA]

def check_quotes(s):
    s = s.replace('"', '')
    s = s.replace("'", "")
    s = s.replace("\t", " ")
    s = s.replace("\n"," ")
    s = s.replace("\r", " ")
    s = s.replace(":", " ")
    s = s.replace("|", " ")
    s = s.replace('Á','A')
    s = s.replace('á','a')
    s = s.replace('É', 'E')
    s = s.replace('é', 'e')
    s = s.replace('Í', 'I')
    s = s.replace('í', 'i')
    s = s.replace('Ó', 'O')
    s = s.replace('ó', 'o')
    s = s.replace('Ú', 'U')
    s = s.replace('ú', 'u')
    s = s.replace('Ñ', 'N')
    s = s.replace('ñ', 'n')
    return s

def transformarANumeroReal(s):
    if s is None:
        return None
    return str(s).replace(",", ".")

def estandarizar(s):
    if s is None:
        return None
    s = str(s).upper()
    return check_quotes(s)

def get_acronimo_nodo(feature):
    prefijo = "NODO:"
    description = feature.getAttribute("kml_description")
    try:
        if description is not None and prefijo in description:
            nodo = description[description.index(prefijo) + len(prefijo):]
            if ' ' in nodo:
                nodo = nodo[:nodo.index(' ')]
            if '\n' in nodo:
                nodo = nodo[:nodo.index('\n')]
            return nodo
    except ValueError:
        print(f"Nodo no encontrado para [{description}]")
    return None

def get_acronimo_sitio(feature):
    description = feature.getAttribute("kml_description")
    if description is not None and ':' in description:
        return description[description.index(':') + 1:description.index('\n')].strip()
    return description

def get_acronimo_cable(feature):
    description = feature.getAttribute("kml_description")
    if description is not None and ':' in description:
        if '\n' in description:
            return description[description.index(':') + 1:description.index('\n')].strip()
        return description[description.index(':') + 1:].strip()
    return None

def get_camara_comment(description):
    try:
        description = description.upper()
        if "LATITUD" in description:
            description = description[:description.index('LATITUD')]
        description = description.strip()
        if description == '':
            return None
        return estandarizar(description.strip().upper())
    except:
        return None

def get_camara_name(name):
    if name is None:
        return "SIN NOMBRE"
    name = name.upper()
    if "LATITUD" in name:
        try:
            name = name[:name.index('LATITUD')]
        except:
            return "SIN NOMBRE"
    return name.strip()

def is_box(name):
    if name is None:
        return False
    return "BOX" in name.upper()

def is_fo(name):
    if name is None:
        return False
    return "C.F.O" in name.upper()

def tiene_rienda(name):
    if name is None:
        return False
    return "C/RIENDA" in name.upper()

def close_files(*files):
    for file in files:
        file.close()

def generate_loaddata(path, *files):
    ffsh = open(path + '\loaddata.sh', 'w')
    for file in files:
        ffsh.write("cat '" + os.path.basename(file.name) + "' | redis-cli --pipe\n")
    ffsh.close()

def get_logs():
    now = datetime.now()  # ahora
    unixtime = int(time.mktime(now.timetuple()))
    return f'{unixtime}..1.'

# obtiene el indice de un elemento en una lista
def get_index_of(array, element, start = 0):
    for i in range(start, len(array)):
        if array[i] == element:
            return i

def get_indexes_of_vertices(cables):
    # creamos la lista de cables
    coord_cables = []
    vertices_cables = []  # feature, inicio, fin
    offset = 0
    for feature in cables:
        coordinates = list(map(lambda x: (x[0], x[1]), feature.getAllCoordinates()))
        coord_cables.extend(coordinates)
        # print(f"[INFO] Longitud temporal del array: {len(coord_cables)}")
        offsetNuevo = offset + len(coordinates)
        vertices_cables.append((feature, offset, offsetNuevo))
        # print(f"Cable entre vertices {offset} y {offsetNuevo}")
        offset = offsetNuevo
    return coord_cables, vertices_cables

def getIndexMasCercano(i, indexes, coord_cables, coord_cajas):
    dist_min = None
    index_real = None

    for j in range(0,len(indexes)):
        index = indexes[j]
        if i == index:
            coord1 = coord_cajas[j]

            coord2 = coord_cables[i]

            dist = cdist([coord1], [coord2])[0][0]

            if dist_min is None or dist_min > dist:
                dist_min = dist
                index_real = j
    return index_real, dist_min

def chequear(i, indexes, coord_cables, coord_cajas):

    index, dist = getIndexMasCercano(i, indexes, coord_cables, coord_cajas)

    if dist is None or dist > 0.001:
        return False
    return True

def getIndexMasCercanoPorTipo(i, indexes, coord_cables, coord_cajas, tipo_cajas, tipo):
    dist_min = None
    index_real = None

    for j in range(0,len(indexes)):
        index = indexes[j]
        if i == index and tipo_cajas[j] == tipo:
            coord1 = coord_cajas[j]

            coord2 = coord_cables[i]

            dist = cdist([coord1], [coord2])[0][0]

            if dist_min is None or dist_min > dist:
                dist_min = dist
                index_real = j
    return index_real, dist_min

def chequear_por_tipo(i, indexes, coord_cables, coord_cajas, tipo_cajas, tipo):
    index, dist = getIndexMasCercanoPorTipo(i, indexes, coord_cables, coord_cajas, tipo_cajas, tipo)

    if dist is None or dist > 0.001:
        return False
    return True

# SI --> True | NO --> False
def transformarEnBooleano(s):
    if s is None:
        return "true"
    if s == "SI":
        return "true"
    if s == "NO":
        return "false"
    return "true"

def transformarAUnixtime(s):
    if s is None:
        return None
    try:
        return str(int(time.mktime(s.timetuple())))
    except AttributeError: # si nos llega mal la fecha, la salteamos
        return None

def obtenerValorRequerido(hoja, fila, columna, valor_por_defecto = "SIN DATO"):
    valor = hoja.cell(row=fila, column=columna).value
    if valor is None:
        return valor_por_defecto
    if type(valor) == str:
        valor = estandarizar(valor)
    return valor

def obtenerDepartamento(s):
    if s is None:
        return "0"
    if s == "25 de Mayo":
        return "2"
    if s == "9 de Julio":
        return "3"
    if s == "Adolfo Alsina":
        return "4"
    if s == "Bariloche":
        return "5"
    if s == "Conesa":
        return "6"
    if s == "General Roca":
        return "7"
    if s == "Pichi Mahuida":
        return "8"
    if s == "Pilcaniyeu":
        return "9"
    if s == "San Antonio":
        return "10"
    if s == "Valcheta":
        return "11"
    if s == "Arauco":
        return "12"
    if s == "Capital":
        return "13"
    if s == "Castro Barros":
        return "14"
    if s == "Chamical":
        return "15"
    if s == "Chilecito":
        return "16"
    if s == "Coronel Felipe Varela":
        return "17"
    if s == "Famatina":
        return "18"
    if s == "Gaiman":
        return "19"
    if s == "General Angel V. Peñaloza":
        return "20"
    if s == "General Belgrano":
        return "21"
    if s == "General Juan Facundo Quiroga":
        return "22"
    if s == "General Lamadrid":
        return "23"
    if s == "General Ocampo":
        return "24"
    if s == "General San Martin":
        return "25"
    if s == "Independencia":
        return "26"
    if s == "Pilagas":
        return "27"
    if s == "Rosario Vera Peñaloza":
        return "28"
    if s == "San Blas de los Andes":
        return "29"
    if s == "Sanagasta":
        return "30"
    if s == "Vinchina":
        return "31"
    if s == "General Alvear":
        return "32"
    if s == "La Paz":
        return "33"
    if s == "Las Heras":
        return "34"
    if s == "Lavalle":
        return "35"
    if s == "Lujan de Cuyo":
        return "36"
    if s == "Malargue":
        return "37"
    if s == "Saladillo":
        return "38"
    if s == "San Carlos":
        return "39"
    if s == "San Rafael":
        return "40"
    if s == "Santa Rosa":
        return "41"
    if s == "Tunuyan":
        return "42"
    if s == "Alumine":
        return "43"
    if s == "Anelo":
        return "44"
    if s == "Catal Lil":
        return "45"
    if s == "Chos Malal":
        return "46"
    if s == "Collon Cura":
        return "47"
    if s == "Confluencia":
        return "48"
    if s == "Gaiman":
        return "49"
    if s == "Huiliches":
        return "50"
    if s == "Lacar":
        return "51"
    if s == "Loncopue":
        return "52"
    if s == "Los Lagos":
        return "53"
    if s == "Minas":
        return "54"
    if s == "Norquin":
        return "55"
    if s == "Pehuenches":
        return "56"
    if s == "Picun Leufu":
        return "57"
    if s == "Picunches":
        return "58"
    if s == "Zapala":
        return "59"
    if s == "Albardon":
        return "62"
    if s == "Calingasta":
        return "63"
    if s == "Caucete":
        return "64"
    if s == "Iglesia":
        return "65"
    if s == "Jachal":
        return "66"
    if s == "San Martin":
        return "67"
    if s == "Santa Lucia":
        return "68"
    if s == "Sarmiento":
        return "69"
    if s == "Ullum":
        return "70"
    if s == "Ambato":
        return "71"
    if s == "Ancasti":
        return "72"
    if s == "Andalgala":
        return "73"
    if s == "Antofagasta de la Sierra":
        return "74"
    if s == "Avellaneda":
        return "75"
    if s == "Belen":
        return "76"
    if s == "Capayan":
        return "77"
    if s == "El Alto":
        return "78"
    if s == "Fray Mamerto Esquiu":
        return "79"
    if s == "La Aguada":
        return "80"
    if s == "La Paz":
        return "81"
    if s == "Las Cañas":
        return "82"
    if s == "Paclin":
        return "83"
    if s == "Poman":
        return "84"
    if s == "Santa Maria":
        return "85"
    if s == "Tinogasta":
        return "86"
    if s == "Valle Viejo":
        return "87"
    if s == "Anta":
        return "88"
    if s == "Cachi":
        return "89"
    if s == "Cafayate":
        return "90"
    if s == "Cerrillos":
        return "91"
    if s == "Chicoana":
        return "92"
    if s == "General Güemes":
        return "93"
    if s == "Guachipas":
        return "94"
    if s == "Iruya":
        return "95"
    if s == "La Caldera":
        return "96"
    if s == "La Candelaria":
        return "97"
    if s == "La Poma":
        return "98"
    if s == "La Viña":
        return "99"
    if s == "Los Andes":
        return "100"
    if s == "Metan":
        return "101"
    if s == "Molinos":
        return "102"
    if s == "Oran":
        return "103"
    if s == "Patiño":
        return "104"
    if s == "Rivadavia":
        return "105"
    if s == "Rosario de la Frontera":
        return "106"
    if s == "Rosario de Lerma":
        return "107"
    if s == "San Carlos":
        return "108"
    if s == "Santa Victoria":
        return "109"
    if s == "Burruyacu":
        return "110"
    if s == "Chicligasta":
        return "111"
    if s == "Cruz Alta":
        return "112"
    if s == "famailla":
        return "113"
    if s == "Graneros":
        return "114"
    if s == "Juan B. Alberdi":
        return "115"
    if s == "La Cocha":
        return "116"
    if s == "Las Cejas":
        return "117"
    if s == "Leales":
        return "118"
    if s == "Lules":
        return "119"
    if s == "Monteros":
        return "120"
    if s == "Rio Chico":
        return "121"
    if s == "Simoca":
        return "122"
    if s == "Tafi del Valle":
        return "123"
    if s == "Tafi Viejo":
        return "124"
    if s == "Trancas":
        return "125"
    if s == "Yerba Buena":
        return "126"
    if s == "Cochinoca":
        return "127"
    if s == "Dr. Manuel Belgrano":
        return "128"
    if s == "El Carmen":
        return "129"
    if s == "General Villegas":
        return "130"
    if s == "Humahuaca":
        return "131"
    if s == "La Almona":
        return "132"
    if s == "Ledesma":
        return "133"
    if s == "Palpala":
        return "134"
    if s == "Rinconada":
        return "135"
    if s == "San Salvador de Jujuy":
        return "136"
    if s == "Santa Barbara":
        return "137"
    if s == "Santa Catalina":
        return "138"
    if s == "Susques":
        return "139"
    if s == "Tilcara":
        return "140"
    if s == "Tumbaya":
        return "141"
    if s == "Valle Grande":
        return "142"
    if s == "Yavi":
        return "143"    
    return "0"

def obtenerTipoSitio(s):
    if s is None:
        return None
    if s == "CASETA":
        return "1"
    if s == "REFEFO":
        return "2"
    if s == "TDA":
        return "3"
    # if s == "Coubicado":
    #     return None
    return None

def obtenerEstadoIngenieria(s):
    if s is None:
        return "SD"
    if s == "Aprobada":
        return "A"
    if s == "Rechazada":
        return "B"
    return "SD"

def obtenerEstadoEnergia(s):
    if s is None:
        return "SD"
    s = s.upper()
    if s == "ENERGIZADO":
        return "E"
    if s == "SIN ENERGIA":
        return "SE"
    return "SD"

# devuelve un sitio por defecto listo para ser insertado
def sitioPorDefecto():
    return {'@oName': "SIN NOMBRE",
            '@identificador': "SIN DATO",
            '@nombre': "SIN DATO",
            '@provincia': obtenerProvincia(PROVINCIA),
            '@departamento': "0",
            '@ing': obtenerEstadoIngenieria(None),
            '@estado': obtenerEstadoGeneral(None),
            '@ene': obtenerEstadoEnergia(None)}

# devuelve la clave del sitio y el diccionario de datos
def obtenerSitio(hoja, fila):
    acronimo = obtenerValorRequerido(hoja, fila, 1)
    nombre = obtenerValorRequerido(hoja, fila, 3)
    sitio = {'@oName': acronimo,
             '@identificador': obtenerValorRequerido(hoja, fila, 2),
             '@nombre': nombre,
             '@provincia': obtenerProvincia(hoja.cell(row=fila, column=4).value),
             '@departamento': obtenerDepartamento(hoja.cell(row=fila, column=5).value),
             '@pob': hoja.cell(row=fila, column=6).value,
             '@proyecto': hoja.cell(row=fila, column=7).value,
             '@lat': transformarANumeroReal(hoja.cell(row=fila, column=8).value),
             '@lon': transformarANumeroReal(hoja.cell(row=fila, column=9).value),
             '@at': transformarAUnixtime(hoja.cell(row=fila, column=10).value),
             '@convenio': transformarAUnixtime(hoja.cell(row=fila, column=11).value),
             '@fecha': transformarAUnixtime(hoja.cell(row=fila, column=12).value),
             '@tipo': obtenerTipoSitio(hoja.cell(row=fila, column=13).value),
             '@ing': obtenerEstadoIngenieria(hoja.cell(row=fila, column=14).value),
             '@estado': obtenerEstadoGeneral(hoja.cell(row=fila, column=28).value),
             '@ene': obtenerEstadoEnergia(hoja.cell(row=fila, column=16).value),
             '@cerco': transformarEnBooleano(hoja.cell(row=fila, column=17).value),
             # mantenimiento
             '@go':hoja.cell(row=fila, column=39).value,
             '@responsable':hoja.cell(row=fila, column=40).value,
             '@sgo':hoja.cell(row=fila, column=41).value,
             '@jefatura':hoja.cell(row=fila, column=42).value,
             '@idMaxRTAA':hoja.cell(row=fila, column=43).value,
             '@redAA':hoja.cell(row=fila, column=44).value,
             '@idMaxRTGE':hoja.cell(row=fila, column=45).value,
             '@refGE':hoja.cell(row=fila, column=46).value,
             '@refFO':hoja.cell(row=fila, column=47).value,
             '@idMaxRG':hoja.cell(row=fila, column=48).value,
             '@refLocal':hoja.cell(row=fila, column=49).value
             }

    agregar_valores_por_defecto(sitio, sitioPorDefecto())
    if acronimo == "SIN DATO":
        if nombre != "SIN DATO":
            return nombre, sitio
    else:
        return acronimo, sitio
    return None, None

def obtenerProyectoNodo(proyecto):
    if proyecto is None:
        return None
    if proyecto == "SH1":
        return "0"
    if proyecto == "SH2":
        return "1"
    return "0"

def obtenerTamanio(tamanio):
    if tamanio is None:
        return "1"
    if tamanio == '3 MTS':
        return "0"
    if tamanio == '6 MTS':
        return "1"
    return "1"

def obtenerTipoShelter(tipo):
    if tipo is None:
        return "4"
    if tipo == "G1":
        return "1"
    if tipo == "G2":
        return "2"
    if tipo == "G3":
        return "3"
    # if tipo == "CH"
    #     return "4"
    return "4"

def obtenerModeloGabinete(s):
    if s is None:
        return "7"
    if s == "AMERINODE":
        return "5"
    if s == "FIBERHOME":
        return "6"
    if s == "Amerinode Outdoor":
        return "1"
    if s == "Amerinode Indoor":
        return "2"
    if s == "Fiberhome Indoor":
        return "3"
    if s == "Fiberhome Outdoor":
        return "4"
    return "7"

def obtenerModeloShelter(s):
    if s is None:
        return "8"
    if s == "SH1-3M":
        return "1"
    if s == "SH1-6M-G1":
        return "2"
    if s == "SH2-3M-G1":
        return "3"
    if s == "SH1-6M-G2":
        return "4"
    if s == "SH2-6M-G1":
        return "5"
    if s == "SH2-6M-G2":
        return "6"
    if s == "SH2-3M-G3":
        return "7"
    return "8" # "SIN DATOS"

def obtenerEstadoGeneral(s):
    if s is None:
        return "E"
    if s == "Construido" or s == "Construccion":
        return "C"
    if s == "Inactivo":
        return "I"
    if s == "Operativo":
        return "E"
    return "E"

def obtenerCuatrimestreEnacom(s):
    if s is None:
        return None
    if s == "Q1":
        return "0"
    if s == "Q2":
        return "1"
    if s == "Q3":
        return "2"
    if s == "Q4":
        return "3"
    if s == "Q5":
        return "4"
    if s == "Q6":
        return "5"
    if s == "Q7":
        return "6"
    if s == "Q8":
        return "7"
    if s == "Q9":
        return "8"
    if s == "Q10":
        return "9"
    if s == "Q11":
        return "10"
    if s == "Q12":
        return "11"
    return None

def obtenerDatosEnacom(hoja, fila, elemento):
    elemento['@comunicado'] = estandarizar(hoja.cell(row = fila, column = 32).value)
    elemento['@fComunicado'] = transformarAUnixtime(hoja.cell(row = fila, column = 33).value)
    elemento['@protocolo'] = estandarizar(hoja.cell(row=fila, column=34).value)
    elemento['@fProtocolo'] = transformarAUnixtime(hoja.cell(row=fila, column=35).value)
    elemento['@utn'] = estandarizar(hoja.cell(row=fila, column=36).value)
    elemento['@fUtnNota'] = transformarAUnixtime(hoja.cell(row=fila, column=37).value)
    elemento['@cuatri'] = obtenerCuatrimestreEnacom(hoja.cell(row=fila, column=38).value)

def agregar_valores_por_defecto(elemento, valores_por_defecto):
    for variable in valores_por_defecto:
        if variable not in elemento or elemento[variable] is None:
            elemento[variable] = valores_por_defecto[variable]

# devuelve un shelter por defecto listo para ser insertado
def shelterPorDefecto():
    return {'@oName': "SIN NOMBRE",
            '@acronimo': "SIN DATO",
            '@tipo': obtenerTipoShelter(None),
            '@tamano': obtenerTamanio(None),
            '@modeloSh': obtenerModeloShelter(None),
            '@estadoSh': obtenerEstadoGeneral(None),
            '@estadoGen': obtenerEstadoGeneral(None)}

def obtenerShelter(hoja, fila, datos_entorno, datos_rtu):
    acronimo = hoja.cell(row=fila, column=19).value
    if acronimo is None:
        return None, None
    shelter = {'@oName': obtenerValorRequerido(hoja, fila, 3),
               '@acronimo': acronimo,
               '@proyecto': obtenerProyectoNodo(hoja.cell(row=fila, column=21).value),
               '@tipo': obtenerTipoShelter(hoja.cell(row=fila, column=24).value),
               '@tamano': obtenerTamanio(hoja.cell(row=fila, column=23).value),
               '@recepcion': transformarAUnixtime(hoja.cell(row=fila, column=25).value),
               '@nroSerie': estandarizar(hoja.cell(row=fila, column=26).value),
               '@modeloSh': obtenerModeloShelter(hoja.cell(row=fila, column=27).value),
               '@proto': transformarAUnixtime(hoja.cell(row=fila, column=28).value),
               '@puesta': transformarAUnixtime(hoja.cell(row=fila, column=29).value),
               '@estadoSh': obtenerEstadoGeneral(hoja.cell(row=fila, column=30).value),
               '@estadoGen': obtenerEstadoGeneral(hoja.cell(row=fila, column=31).value)}
    obtenerDatosEnacom(hoja, fila, shelter)
    if acronimo in datos_entorno:
        shelter.update(datos_entorno[acronimo])
    if acronimo in datos_rtu:
        shelter.update(datos_rtu[acronimo])
    # le agregamos valores por defecto si no tiene
    agregar_valores_por_defecto(shelter, shelterPorDefecto())
    return acronimo, shelter

def gabinetePorDefecto():
    return {'@oName': "SIN NOMBRE",
            '@acronimo': "SIN DATO",
            '@modeloGa': obtenerModeloGabinete(None),
            '@estadoGa': obtenerEstadoGeneral(None),
            '@estadoGen': obtenerEstadoGeneral(None)}

def obtenerGabinete(hoja, fila, datos_entorno):
    acronimo = hoja.cell(row=fila, column=19).value
    print(acronimo)
    if acronimo is None:
        print(f"{acronimo} no lo agrego a la lista")
        return None, None
    gabinete = {'@oName': obtenerValorRequerido(hoja, fila, 3),
                '@acronimo': acronimo,
                '@proyecto': obtenerProyectoNodo(hoja.cell(row=fila, column=21).value),
                '@recepcion': transformarAUnixtime(hoja.cell(row=fila, column=25).value),
                '@nroSerie': estandarizar(hoja.cell(row=fila, column=26).value),
                '@modeloGa': obtenerModeloGabinete(hoja.cell(row=fila, column=27).value),
                '@proto': transformarAUnixtime(hoja.cell(row=fila, column=28).value),
                '@puesta': transformarAUnixtime(hoja.cell(row=fila, column=29).value),
                '@estadoGa': obtenerEstadoGeneral(hoja.cell(row=fila, column=30).value),
                '@estadoGen': obtenerEstadoGeneral(hoja.cell(row=fila, column=31).value)}
    obtenerDatosEnacom(hoja, fila, gabinete)
    if acronimo in datos_entorno:
        gabinete.update(datos_entorno[acronimo])
    # obtenemos valores del gabinete por defecto
    agregar_valores_por_defecto(gabinete, gabinetePorDefecto())
    return acronimo, gabinete

def casetaPorDefecto():
    return {'@oName': "SIN NOMBRE",
            '@acronimo': "SIN DATO",
            '@tipo': obtenerTipoShelter(None),
            '@tamano': obtenerTamanio(None),
            '@modeloCa': obtenerModeloShelter(None),
            '@estadoCa': obtenerEstadoGeneral(None),
            '@estadoGen': obtenerEstadoGeneral(None),
            '@comunicado': ""}

def obtenerCaseta(hoja, fila, datos_rtu):
    acronimo = hoja.cell(row=fila, column=19).value
    if acronimo is None:
        return None, None
    caseta = {'@oName': obtenerValorRequerido(hoja, fila, 3), '@acronimo': acronimo,
              '@proyecto': obtenerProyectoNodo(hoja.cell(row=fila, column=21).value),
              '@tipo': obtenerTipoShelter(hoja.cell(row=fila, column=24).value),
              '@tamano': obtenerTamanio(hoja.cell(row=fila, column=23).value),
              '@recepcion': transformarAUnixtime(hoja.cell(row=fila, column=25).value),
              '@nroSerie': estandarizar(hoja.cell(row=fila, column=26).value),
              '@modeloCa': obtenerModeloShelter(hoja.cell(row=fila, column=27).value),
              '@proto': transformarAUnixtime(hoja.cell(row=fila, column=28).value),
              '@puesta': transformarAUnixtime(hoja.cell(row=fila, column=29).value),
              '@estadoCa': obtenerEstadoGeneral(hoja.cell(row=fila, column=30).value),
              '@estadoGen': obtenerEstadoGeneral(hoja.cell(row=fila, column=31).value)}

    obtenerDatosEnacom(hoja, fila, caseta)
    if acronimo in datos_rtu:
        caseta.update(datos_rtu[acronimo])
    # obtener valores por defecto de la caseta
    agregar_valores_por_defecto(caseta, casetaPorDefecto())
    return acronimo, caseta

def obtenerProvincia(s):
    if s is None:
        return obtenerProvincia(PROVINCIA.title())
    s = s.replace('í','i')
    s = s.replace('á', 'a')
    s = s.replace('ó', 'o')
    s = s.replace('é', 'e')
    s = s.title()
    return s

def obtenerCapacidad(s):
    if s is None:
        return "48"
    if s == "8 F.O":
        return "8"
    if s == "12 F.O":
        return "12"
    if s == "24 F.O":
        return "24"
    if s == "36 F.O":
        return "36"
    if s == "48 F.O":
        return "48"
    if s == "72 F.O":
        return "72"
    if s == "96 F.O":
        return "96"
    if s == "144 F.O":
        return "144"
    if s == "288 F.O":
        return "288"
    return "48"

def obtenerBuffers(s):
    if s is None:
        return "2"
    if s == "1":
        return "0"
    if s == "2":
        return "1"
    if s == "4":
        return "2"
    if s == "6":
        return "3"
    if s == "8":
        return "4"
    if s == "12":
        return "5"
    if s == "24":
        return "6"
    return "2"

def obtenerMarcaCableFO(s):
    if s is None:
        return "9"
    if s == "FURUKAWA":
        return "0"
    if s == "FIBERHOME":
        return "1"
    if s == "PRISMIAN":
        return "2"
    if s == "OPTEL":
        return "3"
    if s == "ZTT":
        return "4"
    if s == "ARTIC":
        return "5"
    if s == "YOFC":
        return "6"
    if s == "4FPRODUCT":
        return "7"
    if s == "MERCURY":
        return "8"
    return "9"

def obtenerTipoCable(s):
    if s is None:
        return "5"
    if s == "OPGW - G652":
        return "0"
    if s == "OPGW - G655":
        return "1"
    if s == "Autosoportado - G652D":
        return "2"
    if s == "Autosoportado - G657":
        return "3"
    if s == "Autosoportado - G655":
        return "4"
    if s == "Ducto - G652D":
        return "5"
    if s == "Ducto - G657":
        return "6"
    if s == "Ducto - G655":
        return "7"
    return "5"

def obtenerLongitud(s):
    if s is None:
        return None
    return s.replace('KM', '')

def obtenerMetodoConstruido(s):
    if s is None:
        return "1"
    if s == "AEREO":
        return "0"
    if s == "CANALIZADO":
        return "1"
    if s == "MIXTO":
        return "2"
    if s == "SUBMARINO":
        return "3"
    return "1"

def obtenerDuctosOcupados(s):
    if s is None:
        return "2"
    if s == "CALLE":
        return "0"
    if s == "CENTRO":
        return "1"
    if s == "MUNICIPAL":
        return "2"
    return "2"

def obtenerTipoRed(s):
    if s is None:
        return "1"
    if s == "TRONCAL":
        return "1"
    if s == "PROVINCIAL":
        return "2"
    if s == "IRU":
        return "3"
    if s == "4":
        return "4"
    return "1"

def obtenerPropietarioCable(s):
    if s is None:
        return "1"
    if s == "ARSAT":
        return "1"
    if s == "ARSAT-SAPEM":
        return "2"
    if s == "SAPEM":
        return "3"
    if s == "ISP":
        return "4"
    if s == "COOPERATIVA":
        return "5"
    if s == "TELCO":
        return "6"
    if s == "ELECTRICA":
        return "7"
    if s == "ASI":
        return "8"
    if s == "OTROS":
        return "9"
    return "1"

def obtenerPropietarioInfraestructura(s):
    if s is None:
        return "1"
    if s == "ARSAT":
        return "1"
    if s == "SAPEM":
        return "2"
    if s == "ASI":
        return "3"
    if s == "TRENES":
        return "4"
    if s == "ISP":
        return "5"
    if s == "COOPERATIVA":
        return "6"
    if s == "TELCO":
        return "7"
    if s == "ELECTRICA":
        return "8"
    if s == "OTROS":
        return "9"
    return "1"

def cablePorDefecto():
    return {'@oName': "SIN NOMBRE",
            '@tramo': "SIN DATO",
            '@prov': obtenerProvincia(None),
            '@estado': 'E',
            '@capacidad': obtenerCapacidad(None),
            '@tipoRed': obtenerTipoRed(None),
            '@buffers': obtenerBuffers(None),
            '@marcaFO': obtenerMarcaCableFO(None),
            '@tipoCable': obtenerTipoCable(None),
            '@datosFO': obtenerTipoCable(None),
            '@metodo': obtenerMetodoConstruido(None),
            '@link': "",
            '@propietarioCable': obtenerPropietarioCable(None),
            '@propietarioInfra': obtenerPropietarioInfraestructura(None),
            '@cao': transformarEnBooleano("SI"),
            '@mediciones': transformarEnBooleano("SI"),
            '@kmz': transformarEnBooleano("SI"),
            '@empalmes': transformarEnBooleano("SI")
    }

def obtenerCable(hoja, fila):
    elemento = {}
    clave = hoja.cell(row=fila, column=2).value
    if clave is None:
        clave = hoja.cell(row=fila, column=3).value
        if clave is None:
            clave = hoja.cell(row=fila, column=4).value
            if clave is None:
                return None
    acronimo = hoja.cell(row=fila, column=2).value
    if acronimo is None:
        acronimo = "SIN DATO"
    elemento['@oName'] = acronimo
    tramo = hoja.cell(row=fila, column=3).value
    if tramo is None:
        tramo = "SIN DATO"
    elemento['@tramo'] = tramo
    elemento['@subtramo'] = hoja.cell(row=fila, column=4).value
    elemento['@prov'] = obtenerProvincia(hoja.cell(row=fila, column=5).value)
    elemento['@estado'] = 'E'
    elemento['@capacidad'] = obtenerCapacidad(hoja.cell(row=fila, column=6).value)
    elemento['@buffers'] = obtenerBuffers(hoja.cell(row=fila, column=7).value)
    elemento['@marcaFO'] = obtenerMarcaCableFO(hoja.cell(row=fila, column=8).value)
    elemento['@tipoCable'] = obtenerTipoCable(hoja.cell(row=fila, column=9).value)
    elemento['@longitud'] = obtenerLongitud(hoja.cell(row=fila, column=10).value)
    elemento['@longitudO'] = obtenerLongitud(hoja.cell(row=fila, column=11).value)
    elemento['@metodo'] = obtenerMetodoConstruido(hoja.cell(row=fila, column=12).value)
    elemento['@ductosO'] = obtenerDuctosOcupados(hoja.cell(row=fila, column=13).value)
    elemento['@tipoRed'] = obtenerTipoRed(hoja.cell(row=fila, column=14).value)
    elemento['@propietarioCable'] = obtenerPropietarioCable(hoja.cell(row=fila, column=15).value)
    elemento['@propietarioInfra'] = obtenerPropietarioInfraestructura(hoja.cell(row=fila, column=16).value)
    elemento['@cao'] = transformarEnBooleano(hoja.cell(row=fila, column=17).value)
    elemento['@mediciones'] = transformarEnBooleano(hoja.cell(row=fila, column=18).value)
    elemento['@kmz'] = transformarEnBooleano(hoja.cell(row=fila, column=19).value)
    elemento['@empalmes'] = transformarEnBooleano(hoja.cell(row=fila, column=20).value)
    link = hoja.cell(row=fila, column=21).value
    if link is None:
        link = "SIN DATO"
    elemento['@link'] = link
    elemento['@rnfo'] = hoja.cell(row=fila, column=22).value
    elemento['@cgo'] = hoja.cell(row=fila, column=23).value
    elemento['@go'] = hoja.cell(row=fila, column=24).value
    elemento['@contratista'] = hoja.cell(row=fila, column=25).value
    # obtenemos los valores por defecto
    agregar_valores_por_defecto(elemento, cablePorDefecto())
    return clave, elemento

def poblar_sidx(val):
    sidx = {}
    if '@oName' in val:
        name = val['@oName']
        if name is not None and name != '':
            sidx["@oName"] = name
    if '@acronimo' in val:
        acronimo = val['@acronimo']
        if acronimo is not None and acronimo != '':
            sidx["@acronimo"] = acronimo
    if '@nombre' in val:
        nombre = val['@nombre']
        if nombre is not None and nombre != '':
            sidx["@nombre"] = nombre
    if '@kmlId' in val:
        kmlId = val['@kmlId']
        if kmlId is not None:
            sidx['@kmlId'] = kmlId
    if '@subtramo' in val and val['@subtramo'] is not None:
        sidx['@subtramo'] = val['@subtramo']
    if '@tramo' in val and val['@tramo'] != 'SIN DATO':
        sidx['@tramo'] = val['@tramo']
    return sidx

def obtenerProvinciaEntorno(s):
    if s is None:
        return None
    return PROVINCIAS_ENTORNO[s]

def obtenerPotencia(s):
    if s is None:
        return "3"
    s = str(s)
    if s == "45":
        return "1"
    if s == "100":
        return "2"
    return "3"

def obtenerMarcaRTU(s):
    if s is None:
        return "2"
    if s == "EXFO":
        return "1"
    return "2"

def obtenerModeloRTU(s):
    if s is None:
        return "3"
    if s == "FG-700":
        return "1"
    if s == "FG-720":
        return "2"
    return "3"

def move_coordinates(coordinates):
    lat = coordinates[0]
    long = coordinates[1]
    return lat + 0.0005, long + 0.0005

class FeatureProcessor(object):

    # declaramos todas las variables de clase
    def __init__(self):
        self.featureList = []
        self.folderList = {} # indexado por kml_id
        self.styleList = []
        self.styleMapList = []
        self.ffo = None
        self.ffocfg = None
        self.ffval = None
        self.ffsidx = None
        self.ffv = None
        self.ffgeoidx = None
        self.ffio = None
        self.ffco = None
        self.object_id = OBJECT_ID_INICIAL
        self.company_id = 30
        self.logs = get_logs()
        self.cables = []
        self.tipo_cable = []
        self.cajas_coord = [] #  coordenadas
        self.cajas_id = [] # ids
        self.tipo_camara = []
        self.empalme_por_caja = {}
        self.camara_por_caja = {} # id de caja --> id de camara
        self.cajas_de_empalme = []
        self.fronteras = []
        self.cajas_frontera = [] # ids de cajas
        # datos del excel
        self.tramos_y_subtramos = {}
        self.sitios = {}
        self.shelters = {}
        self.gabinetes = {}
        self.casetas = {}
        self.inactivos = []
        self.datos_entorno = {}
        self.datos_rtu = {}

    def actualizarIdMaximo(self, ffpath):
        with open(ffpath + "/MAX_OID.txt", "w") as f:
            # actualizamos el oid maximo
            f.write(f'SET {self.company_id}.MAX_OID "{self.object_id + 1}"\n')

    def poblarTramosYSubtramos(self):
        wb = load_workbook(filename=TRAMOS_Y_SUBTRAMOS_EXCEL_PATH)
        hoja = wb.active
        for i in range(4,hoja.max_row + 1):
            provincia = hoja.cell(row=i, column=5).value
            if provincia is None:
                continue
            # diccionario --> indexado por acronimo --> elemento
            clave, cable = obtenerCable(hoja, i)
            if clave is None:
                continue
            self.tramos_y_subtramos[clave] = cable
        wb.close()

    def poblarDatosEntorno(self):
        wb = load_workbook(filename=ENTORNO_EXCEL_PATH)
        hoja = wb.active
        for i in range(2, hoja.max_row + 1):
            elemento = {}
            # obtenemos la provincia, si no es la que estamos migrando la salteamos
            provincia = obtenerProvinciaEntorno(hoja.cell(i,3).value)
            if provincia != PROVINCIA:
                continue
            acronimo = hoja.cell(i,1).value
            if acronimo is None:
                continue
            elemento["@bateria"] = estandarizar(hoja.cell(i,5).value)
            elemento["@potencia"] = obtenerPotencia(hoja.cell(i,6).value)
            elemento["@alarma"] = estandarizar(hoja.cell(i,7).value)
            elemento["@ip"] = hoja.cell(i,8).value
            elemento["@ipWE"] = hoja.cell(i,9).value
            elemento["@ipWI"] = hoja.cell(i,10).value
            elemento["@ipGE"] = hoja.cell(i,11).value
            elemento["@ipNVR"] = hoja.cell(i,12).value
            elemento["@ipR"] = hoja.cell(i,13).value
            elemento["@ipTTA"] = hoja.cell(i,14).value
            elemento["@ipFC"] = hoja.cell(i,15).value
            elemento["@grupoElectrogeno"] = "true"
            elemento["@marca"] = "1"
            elemento["@modelo"] = "1"
            self.datos_entorno[acronimo] = elemento
        wb.close()

    def poblarDatosRTU(self):
        wb = load_workbook(filename=RTU_EXCEL_PATH)

        hoja = wb.active

        for i in range(4,hoja.max_row + 1):

            provincia = hoja.cell(i, 4).value
            acronimoNodo = estandarizar(hoja.cell(i, 2).value)

            if provincia is None or provincia != PROVINCIA or acronimoNodo is None:
                continue

            if acronimoNodo in self.datos_rtu:
                rtu = self.datos_rtu[acronimoNodo]
                tramo1 = estandarizar(hoja.cell(i, 15).value)
                if tramo1 is not None:
                    rtu["@ruta1"] = rtu["@ruta1"] + " / " + estandarizar(hoja.cell(i, 15).value)
                tramo2 = estandarizar(hoja.cell(i, 19).value)
                if tramo2 is not None:
                    rtu["@ruta2"] = rtu["@ruta2"] + " / " + estandarizar(tramo2)
                tramo3 = estandarizar(hoja.cell(i, 23).value)
                if tramo3 is not None:
                    rtu["@ruta3"] = rtu["@ruta3"] + " / " + estandarizar(tramo3)
                tramo4 = estandarizar(hoja.cell(i, 27).value)
                if tramo4 is not None:
                    rtu["@ruta4"] = rtu["@ruta4"] + " / " + estandarizar(tramo4)
                tramo5 = estandarizar(hoja.cell(i, 31).value)
                if tramo5 is not None:
                    rtu["@ruta5"] = rtu["@ruta5"] + " / " + estandarizar(tramo5)
                tramo6 = estandarizar(hoja.cell(i, 35).value)
                if tramo6 is not None:
                    rtu["@ruta6"] = rtu["@ruta6"] + " / " + estandarizar(tramo6)
                tramo7 = estandarizar(hoja.cell(i, 39).value)
                if tramo7 is not None:
                    rtu["@ruta7"] = rtu["@ruta7"] + " / " + estandarizar(tramo7)
                tramo8 = estandarizar(hoja.cell(i, 43).value)
                if tramo8 is not None:
                    rtu["@ruta8"] = rtu["@ruta8"] + " / " + estandarizar(tramo8)

                self.datos_rtu[acronimoNodo] = rtu

                continue

            element = {
                "@id": estandarizar(hoja.cell(i, 5).value),
                "@marca": obtenerMarcaRTU(hoja.cell(i, 6).value),
                "@modelo": obtenerModeloRTU(hoja.cell(i, 7).value),
                "@puertos": estandarizar(hoja.cell(i, 9).value),
                "@pPuerto1": estandarizar(hoja.cell(i, 14).value),
                "@ruta1": estandarizar(hoja.cell(i, 15).value),
                "@pPuerto2": estandarizar(hoja.cell(i, 18).value),
                "@ruta2": estandarizar(hoja.cell(i, 19).value),
                "@pPuerto3": estandarizar(hoja.cell(i, 22).value),
                "@ruta3": estandarizar(hoja.cell(i, 23).value),
                "@pPuerto4": estandarizar(hoja.cell(i, 26).value),
                "@ruta4": estandarizar(hoja.cell(i, 27).value),
                "@pPuerto5": estandarizar(hoja.cell(i, 30).value),
                "@ruta5": estandarizar(hoja.cell(i, 31).value),
                "@pPuerto6": estandarizar(hoja.cell(i, 34).value),
                "@ruta6": estandarizar(hoja.cell(i, 35).value),
                "@pPuerto7": estandarizar(hoja.cell(i, 38).value),
                "@ruta7": estandarizar(hoja.cell(i, 39).value),
                "@pPuerto8": estandarizar(hoja.cell(i, 42).value),
                "@ruta8": estandarizar(hoja.cell(i, 43).value)
            }

            self.datos_rtu[acronimoNodo] = element

        wb.close()

    def poblarSitiosYNodos(self):
        wb = load_workbook(filename=SITIOS_EXCEL_PATH)

        hoja = wb.active

        for i in range(3,hoja.max_row + 1):
            estado = hoja.cell(row=i, column=31).value
            # si es inactivo lo salteamos pero agregamos el acronimo a la planilla de inactivos
            if estado == "Inactivo":
                self.inactivos.append(hoja.cell(row=i, column=1).value)
                print(f"INACTIVO: ({i}) Acronimo {hoja.cell(row=i, column=1).value} Nombre {hoja.cell(row=i, column=3).value}")
                continue
            # guardamos el sitio
            clave_sitio, sitio = obtenerSitio(hoja = hoja, fila = i)
            if clave_sitio is not None:
                self.sitios[clave_sitio] = sitio
            # guardamos el nodo
            tipo = hoja.cell(row=i, column=18).value
            # si el tipo es None lo salteamos
            if tipo is None:
                continue
            if tipo == "Shelter":
                clave, nodo = obtenerShelter(hoja, i, self.datos_entorno, self.datos_rtu)
                self.shelters[clave] = nodo
            elif tipo == "Gabinete" or tipo == "Coubicado":
                clave, nodo = obtenerGabinete(hoja, i, self.datos_entorno)
                self.gabinetes[clave] = nodo
            elif tipo == "Caseta":
                clave, nodo = obtenerCaseta(hoja, i, self.datos_rtu)
                self.casetas[clave] = nodo
        wb.close()

        print("Elementos tanto activos como inactivos: ")
        for inactivo in self.inactivos:
            if inactivo in self.sitios:
                print(f"SITIO {inactivo}")
            elif inactivo in self.shelters:
                print(f"SHELTER {inactivo}")
            elif inactivo in self.gabinetes:
                print(f"GABINETE {inactivo}")

    def input(self, feature):
        fme_feature_type = feature.getAttribute("fme_feature_type")
        if fme_feature_type in ELEMENT_BASENAMES:
            self.featureList.append(feature)
        elif fme_feature_type == "Folder" or fme_feature_type == "Document":
            kml_id = feature.getAttribute("kml_id")
            self.folderList[kml_id] = feature
        elif fme_feature_type == "Style":
            self.styleList.append(feature)
        elif fme_feature_type == "StyleMap":
            self.styleMapList.append(feature)

    # gets the folder (camara, fo, red, sitios, etc)
    def get_folder(self,feature):
        try:
            father = self.folderList[feature.getAttribute("kml_parent")]
            filename = father.getAttribute("kml_name")
            if filename in TIPOS_SITIO or filename in TRAMO or filename in SUBTRAMO: ##if it is, return it
                return filename
            else:
                return self.get_folder(father) ##if not, repeat with the father
        except (ValueError, KeyError):
            return None ##if it is the root (ValueError) or it doesn't have a father (KeyError), None

    # gets the folder (camara, fo, red, sitios, etc)
    def get_root_folder(self,feature):
        try:
            father = self.folderList[feature.getAttribute("kml_parent")]
            filename = father.getAttribute("kml_name")
            if filename in CARPETAS_PRINCIPALES: ##if it is, return it
                return filename
            else:
                return self.get_root_folder(father) ##if not, repeat with the father
        except (ValueError, KeyError):
            return None ##if it is the root (ValueError) or it doesn't have a father (KeyError), None

    # network own_library: 1 == troncal, 2 == distribucion, 3 == clientes, 4 == infraestructura, 5 == areas de zonas
    def get_ocfg(self,feature, folder, name, f_type):
        name = estandarizar(name)
        if f_type == "fme_collection":  ##if it is a collection, skip it
            return None
        if folder in TRITUBO:
            return "gc/duc", {
                "@oName": name,
                "@tendido": "1",
                "@ocupacion1": "true",
                "@nc1": name,
                "@propietarioTri": "1",
                "@estado": "E"
            }, obtener_network_id_infra()
        # si es una torre, no nos interesa
        if f_type == "fme_line" and "TORRE" in name.upper():
            return None
        # tenemos que ver si matchea con el excel o si tiene acronimo
        # en definitiva siempre que tenga acronimo se matchea
        if folder in TRAMO or folder in SUBTRAMO:
            self.cables.append(feature)
            self.tipo_cable.append(folder)
            return None
        if folder in SITIO:
            acronimo = estandarizar(get_acronimo_sitio(feature))
            if (acronimo in self.inactivos and acronimo not in self.sitios) or acronimo not in self.sitios:
                print(f"[INFO] SITIO EN KMZ {acronimo} NO MATCHEA")
                return None
            val = self.sitios[acronimo]
            self.sitios.pop(acronimo)
            print(f"[INFO] Escribiendo SITIO '{acronimo}'")
            return "go/fo/si", val, obtener_network_id_infra()
        if folder in SHELTER:
            acronimo = estandarizar(get_acronimo_nodo(feature))
            if acronimo in self.shelters:
                val = self.shelters[acronimo]
            else:
                val = shelterPorDefecto()
                if name is not None:
                    val["@oName"] = name
                if acronimo is not None:
                    val["@acronimo"] = acronimo
            if val["@tamano"] == "0":
                return "go/fo/sh3", val, obtener_network_id_fo()
            return "go/fo/sh", val, obtener_network_id_fo()
        if folder in GABINETE:
            acronimo = estandarizar(get_acronimo_nodo(feature))
            if acronimo in self.gabinetes:
                val = self.gabinetes[acronimo]
                return "go/fo/gb", val, obtener_network_id_fo()
            return None
        if folder in CASETA:
            acronimo = estandarizar(get_acronimo_nodo(feature))
            if acronimo in self.casetas:
                val = self.casetas[acronimo]
                return "go/fo/cs", val, obtener_network_id_infra()
            return None
        if folder in CAMARA:
            if is_box(name):
                tipo_empalme = "2"
                descripcion = feature.getAttribute("kml_description")
                if descripcion is not None and "DERIVACION" in descripcion or "C.D" in name:
                    tipo_empalme = "1"
                val = {
                    "@oName": name,
                    "@tipoEmpalme":tipo_empalme,
                    "@tipoCaja": "5",
                    "@estadoEm": "E",
                    "@ocupacion": "1"
                }
                if is_fo(name):
                    val["@frontera"] = "true"
                return "go/fo/em", val, obtener_network_id_fo()
            if name is None:
                name = "SIN NOMBRE"
            val = {
                "@oName":estandarizar(name),
                "@aCaja":"SIN DATOS",
                "@tipoCam":"3",
                "@tapas":"3",
                "@estadoCam":"0",
                "@material":"1",
                "@propietarioCam":"1",
                "@estado":"E",
                "@marker":"true"
            }
            return "go/fo/cam", val, obtener_network_id_infra()
        if folder in EDT:
            return "go/fo/edts", {
                "@oName": name
            }, obtener_network_id_fo()
        if folder in POSTE:
            val = {
                "@oName": name,
                "@material": "1",
                "@altura": "2",
                "@tipoPo": "T",
                "@rienda": "false",
                "@propietarioPos": "1",
                "@estadoPo": "0",
                "@estado": "E"
            }
            if tiene_rienda(name):
                val["@rienda"] = "true"
            return "go/fo/pos", val, obtener_network_id_infra()
        return None

    def write_object(self, network_id, ocfg, vals, sidxs, coordinates):
        self.object_id += 1
        key = f'{self.company_id}.{network_id}.{self.object_id}'
        # SET 10.1.1 "0..1."
        self.ffo.write(f'SET {key} "{self.logs}"\n')
        # SADD 10.1.1:ocfg "0..1.:a/cgc/a"
        # print(f'SADD {key}:ocfg "{self.logs}:{ocfg}"\n')
        self.ffocfg.write(f'SADD {key}:ocfg "{self.logs}:{ocfg}"\n')

        for variable in vals:
            valor = vals[variable]
            if valor is None:
                continue
            valor = check_quotes(str(valor))
            # SADD 10.1.1:val "0..1.:@oName|CAMPO INDIO|0"
            self.ffval.write(f'SADD {key}:val "{self.logs}:{variable}|{valor}|0"\n')

        for variable in sidxs:
            valor = sidxs[variable]
            if valor is None:
                continue
            valor = check_quotes(valor)
            self.ffsidx.write(f'ZADD {self.company_id}.{variable}.sidx 0 "{valor}:{self.logs}:{key}"\n')

        geoidx_id = f'{self.company_id}.{network_id}'

        # write vertex
        n = len(coordinates)
        i = 0
        while i < n:
            long, lat = coordinates[i]
            s_lat = "%.13f" % lat
            s_lon = "%.13f" % long
            # SADD 10.1.1:v "0..1.:-50.7690301903|-70.7467998635|0|0"
            self.ffv.write(f'SADD {key}:v "{self.logs}:{s_lat}|{s_lon}|{i}|0"\n')
            # GEOADD 10.1:geoidx -70.7467998635 -50.7690301903 "0..1.:10.1.1|0|9"
            self.ffgeoidx.write(f'GEOADD {geoidx_id}:geoidx {s_lon} {s_lat} "{self.logs}:{key}|{i}|{n}"\n')
            i += 1
        return key

    # returns a dictionary with the io objects buffercolor -> fibercolor -> id
    def create_fo_ios(self, fo_id, buffers_num, total_fibers_num):
        fibers_num = int(total_fibers_num / buffers_num)
        dicc = {} # buffercolor --> fibercolor --> id
        for i in range(0, buffers_num):
            vals = {"@io" : fo_id,
                    "@io0" : fo_id,
                    "@color" : buffer_colors[i],
                    "@name" : buffer_colors[i],
                    "@order" : i
                    }
            buff_str_id = self.write_object(100, "io/fo/t", vals, {} , [])
            self.ffio.write(f'SADD {fo_id}:io "{self.logs}:{buff_str_id}"\n')
            dicc[buffer_colors[i]] = {} # color -> empalme
            for j in range(0, fibers_num):
                vals = {"@io": buff_str_id,
                        "@io0": buff_str_id,
                        "@color": fiber_colors[j],
                        "@name": f'{buffer_colors[i]}-{fiber_colors[j]}',
                        "@order": j
                        }
                fib_str_id = self.write_object(100, "io/fo/h", vals, {}, [])
                self.ffio.write(f'SADD {buff_str_id}:io "{self.logs}:{fib_str_id}"\n')
                dicc[buffer_colors[i]][fiber_colors[j]] = fib_str_id
        return dicc

    def conectar_io(self, caja_id, io):

        # parametrizamos los numeros de los extremos de la conexion
        # TODO alternar entre 1 y 2 hilos y empalmes
        n1 = 2
        n2 = 1
        if caja_id in self.empalme_por_caja:
            for buffercolor in io:
                for fibercolor in io[buffercolor]:
                    hilo = io[buffercolor][fibercolor]
                    empalme_id = self.empalme_por_caja[caja_id][buffercolor][fibercolor]
                    self.ffio.write(f'SADD {caja_id}:io "{self.logs}:{empalme_id}"\n')
                    self.ffco.write(f'SADD {hilo}:co "{self.logs}:{n1}|{empalme_id}|{n2}"\n')
                    self.ffco.write(f'SADD {empalme_id}:co "{self.logs}:{n2}|{hilo}|{n1}"\n')
        else:
            empalmes = {}
            val = {"@io": caja_id,
                   "@io0": caja_id,
                   "@sType": "fusion",
                   "@sName": "",
                   "@spliceFrom": "OP"}

            for buffercolor in io:
                for fibercolor in io[buffercolor]:
                    hilo = io[buffercolor][fibercolor]
                    empalme_id = self.write_object(100, "io/fo/e", val, {}, [])
                    if buffercolor not in empalmes:
                        empalmes[buffercolor] = {}
                    empalmes[buffercolor][fibercolor] = empalme_id
                    self.ffio.write(f'SADD {caja_id}:io "{self.logs}:{empalme_id}"\n')
                    self.ffco.write(f'SADD {hilo}:co "{self.logs}:{n1}|{empalme_id}|{n2}"\n')
                    self.ffco.write(f'SADD {empalme_id}:co "{self.logs}:{n2}|{hilo}|{n1}"\n')
            self.empalme_por_caja[caja_id] = empalmes



    def connect(self, key, other_key, n1, n2):
        self.ffco.write(f'SADD {key}:co "{self.logs}:{n1}|{other_key}|{n2}"\n')
        self.ffco.write(f'SADD {other_key}:co "{self.logs}:{n2}|{key}|{n1}"\n')

    def divide_cable(self):

        coord_cajas_sin_co = self.cajas_coord.copy()
        cajas_id_sin_co = self.cajas_id.copy()
        cajas_tipo_sin_co = self.tipo_camara.copy()

        cables_creados = []
        tritubos_creados = []

        coord_cables, vertices_cables = get_indexes_of_vertices(self.cables)

        tree_cables = KDTree(coord_cables)
        _, indexes = tree_cables.query(self.cajas_coord,k = 1)

        print(f"[INFO] Cantidad de cajas de empalme: {len(self.camara_por_caja)}")
        print(f"[INFO] Cantidad de camaras: {len(indexes)}")

        contador_cables_totales = 0
        contador_tritubos_totales = 0
        index_cable = 0
        # iteramos sobre todos los cables
        for feature, inicio, fin in vertices_cables:
            print(f"[INFO] IN, FIN: {inicio}, {fin}")
            print(f"[INFO] CABLE: {feature.getAttribute('kml_name')}")
            vertice_inicial = None
            coord_iniciales = None
            id_inicial = None
            tritubos_creados.append([])

            tipo_cable = self.tipo_cable[index_cable]

            tri_vertice_inicial = None
            tri_coord_iniciales = None
            tri_id_inicial = None
            contador = 0
            tri_contador = 0
            # iteramos sobre todos los vertices del cable
            # para el primer vertice, si no esta conectado a nada hay que ponerlo como primer vertice

            if not chequear(inicio, indexes, coord_cables, self.cajas_coord):

                vertice_inicial = inicio
                coord_iniciales = coord_cables[inicio]
                tri_vertice_inicial = vertice_inicial
                tri_coord_iniciales = coord_iniciales

            for i in range(inicio,fin):
                if chequear_por_tipo(i, indexes, coord_cables, self.cajas_coord, self.tipo_camara, "go/fo/em"):
                    index_caja, _ = getIndexMasCercanoPorTipo(i, indexes, coord_cables, self.cajas_coord,self.tipo_camara, "go/fo/em")
                    coord_final = self.cajas_coord[index_caja]
                    coord_cables[i] = coord_final
                    caja_id = self.cajas_id[index_caja]

                    # print(f"[DEBUG] Vertice {i - inicio} chequea contra caja {caja_id}")
                    camara_id = self.camara_por_caja[caja_id]

                    # no los puedo borrar ya que esto afectaria los indices del resto de la lista
                    cajas_id_sin_co[index_caja] = None
                    cajas_tipo_sin_co[index_caja] = None
                    coord_cajas_sin_co[index_caja] = None

                    vertice = i

                    # si es el primer vertice, inicializamos nada mas
                    if vertice_inicial is None:
                        vertice_inicial = vertice
                        coord_iniciales = coord_final
                        id_inicial = caja_id
                    else:
                        print(f"[INFO] Creando cable entre los vertices: {vertice_inicial - inicio} : {vertice - inicio}")
                        contador += 1
                        # las coordenadas seran la coordenada del cable inicial, la del coord_final y los de en medio
                        coordinates = [coord_iniciales]
                        coordinates.extend(coord_cables[vertice_inicial + 1: vertice])
                        coordinates.append(coord_final)
                        cables_creados.append( (coordinates, (id_inicial, caja_id, feature, tipo_cable)) )

                        vertice_inicial = vertice
                        coord_iniciales = coord_final
                        id_inicial = caja_id

                    # tambien manejamos el tritubo
                    if tri_vertice_inicial is None:
                        tri_vertice_inicial = vertice
                        tri_coord_iniciales = coord_final
                        tri_id_inicial = camara_id
                    else:
                        print(f"[INFO] Creando tritubo entre los vertices: {tri_vertice_inicial- inicio} : {vertice- inicio}")
                        tri_contador += 1
                        # las coordenadas seran la coordenada del cable inicial, la del coord_final y los de en medio
                        coordinates = [tri_coord_iniciales]
                        coordinates.extend(coord_cables[tri_vertice_inicial + 1: vertice])
                        coordinates.append(coord_final)
                        tritubos_creados[-1].append( (coordinates, (tri_id_inicial, camara_id)) )
                        tritubos_creados.append([])
                        tri_vertice_inicial = vertice
                        tri_coord_iniciales = coord_final
                        tri_id_inicial = camara_id

                # si el vertice esta conectado a una caja
                elif chequear(i,indexes, coord_cables, self.cajas_coord):
                    index_caja, _ = getIndexMasCercano(i, indexes, coord_cables, self.cajas_coord)
                    coord_final = self.cajas_coord[index_caja]
                    coord_cables[i] = coord_final

                    tipo = self.tipo_camara[index_caja]
                    # si es una camara, tenemos que manejar el tritubo
                    # TODO este IF esta de mas, aca siempre van a llegar camaras
                    if tipo == "go/fo/cam":
                        camara_id = self.cajas_id[index_caja]

                        # print(f"[DEBUG] Vertice {i - inicio} chequea contra camara {camara_id}")
                        vertice = i

                        # si es el primer vertice, inicializamos el cable
                        if vertice_inicial is None:
                            vertice_inicial = vertice
                            coord_iniciales = coord_final
                        # si es el primer vertice, inicializamos nada mas
                        if tri_vertice_inicial is None:
                            tri_vertice_inicial = vertice
                            tri_coord_iniciales = coord_final
                            tri_id_inicial = camara_id
                            continue
                        print(f"[INFO] Creando tritubo entre los vertices: {tri_vertice_inicial - inicio} : {vertice - inicio}")
                        tri_contador += 1
                        # las coordenadas seran la coordenada del cable inicial, la del coord_final y los de en medio
                        coordinates = [tri_coord_iniciales]
                        coordinates.extend(coord_cables[tri_vertice_inicial + 1: vertice])
                        coordinates.append(coord_final)
                        tritubos_creados[-1].append( (coordinates, (tri_id_inicial, camara_id)) )

                        # no los puedo borrar ya que esto afectaria los indices del resto de la lista
                        cajas_id_sin_co[index_caja] = None
                        cajas_tipo_sin_co[index_caja] = None
                        coord_cajas_sin_co[index_caja] = None

                        tri_vertice_inicial = vertice
                        tri_coord_iniciales = coord_final
                        tri_id_inicial = camara_id
                        continue
            # para el cable usamos esta funcion porque si hay un match con una camarano se va a cortar,
            # por lo tanto preguntamos si hay algun matcheo exclusivamente con una caja de empalmes
            if vertice_inicial != fin - 1:
                # si es el primer vertice, inicializamos nada mas
                if vertice_inicial is None:
                    vertice_inicial = inicio
                    coord_iniciales = coord_cables[inicio]
                print(f"[INFO] Creando cable entre los vertices: {vertice_inicial- inicio} : {fin - inicio- 1}")
                contador += 1
                # las coordenadas seran la coordenada del cable inicial, la del coord_final y los de en medio
                coordinates = [coord_iniciales]
                coordinates.extend(coord_cables[vertice_inicial + 1: fin])
                cables_creados.append( (coordinates, (id_inicial, None, feature, tipo_cable)) )
            # para el tritubo usamos esta funcion porque si es una camara o una caja se va a cortar igual,
            # por lo tanto preguntamos si hay algun matcheo con algun elemento y es suficiente
            if tri_vertice_inicial != fin - 1:
                # lo mismo para el tritubo
                if tri_vertice_inicial is None:
                    tri_vertice_inicial = inicio
                    tri_coord_iniciales = coord_cables[inicio]
                print(f"[INFO] Creando tritubo entre los vertices: {tri_vertice_inicial- inicio} : {fin - inicio- 1}")
                tri_contador += 1
                # las coordenadas seran la coordenada del cable inicial, la del coord_final y los de en medio
                coordinates = [tri_coord_iniciales]
                coordinates.extend(coord_cables[tri_vertice_inicial + 1: fin])

                tritubos_creados[-1].append( (coordinates, (tri_id_inicial, None)) )
            if not tritubos_creados[-1]: # si el ultimo elemento es []
                tritubos_creados.pop(-1)
            print(f"[INFO] Cantidad de cables: {contador}")
            print(f"[INFO] Cantidad de tritubos: {tri_contador}")
            if contador == 0 or tri_contador == 0:
                print("[ERROR] Error generando cable!")
                exit(-1)
            if len(cables_creados) != len(tritubos_creados):
                print("[ERROR] ERROR HACIENDO LISTAS DE CABLES Y TRITUBOS CREADOS!")
                print(cables_creados)
                print(tritubos_creados)
                print(feature.getAttribute("kml_name"))
                exit(-1)
            index_cable+=1
            contador_cables_totales += contador
            contador_tritubos_totales += tri_contador
        print(f"[INFO] Total de cables generados: {contador_cables_totales}")
        print(f"[INFO] Total de tritubos generados: {contador_tritubos_totales}")
        cajas_id_sin_co = list(filter(lambda x: x is not None, cajas_id_sin_co))
        cajas_tipo_sin_co = list(filter(lambda x: x is not None, cajas_tipo_sin_co))
        coord_cajas_sin_co = list(filter(lambda x: x is not None, coord_cajas_sin_co))
        AgregadorDeVertices.corregir_cables(coord_cajas_sin_co,
                                            cajas_id_sin_co,
                                            cajas_tipo_sin_co,
                                            self.camara_por_caja,
                                            cables_creados,
                                            tritubos_creados)
        self.crear_cables_y_tritubos(cables_creados, tritubos_creados)

    def crear_cables_y_tritubos(self, cables, tritubos):

        contador_tritubos = 0
        for indice_cable in range(0, len(cables)):
            conexion_inicial, conexion_final, feature, tipo_cable = cables[indice_cable][1]

            # todos los cables que lleguen hasta aca van a tener acronimo,
            # los que no tienen los filtramos antes
            acronimo = get_acronimo_cable(feature)
            if acronimo is not None and acronimo in self.tramos_y_subtramos:
                vals_cable = self.tramos_y_subtramos[acronimo].copy()
            else: # por defecto
                vals_cable = cablePorDefecto()
                if acronimo is None:
                    acronimo = "SIN NOMBRE"
                vals_cable["@oName"] = acronimo
                name = feature.getAttribute("kml_name")
                if name is None:
                    name = "SIN NOMBRE"
                if tipo_cable in TRAMO:
                    vals_cable["@tramo"] = estandarizar(name)
                else:
                    vals_cable["@tramo"] = "SIN DATO"
                    vals_cable["@subtramo"] = estandarizar(name)
            if conexion_inicial is not None:
                vals_cable["@foSide-" + conexion_inicial] = 'OUT'
            if conexion_final is not None:
                vals_cable["@foSide-" + conexion_final] = 'IN'
            sidx_cable = poblar_sidx(vals_cable)


            coordenadas_cable = cables[indice_cable][0]
            # escribimos
            cable_key = self.write_object(obtener_network_id_fo(), "gc/fo", vals_cable, sidx_cable, coordenadas_cable)
            cable_io = self.create_fo_ios(fo_id=cable_key,
                                          buffers_num=BUFFERS_PER_TYPE[vals_cable["@buffers"]],
                                          total_fibers_num=FIBER_PER_TYPE[vals_cable["@capacidad"]])
            if conexion_inicial is not None:
                self.connect(cable_key, conexion_inicial, n1 = 1, n2 = 1)
                self.conectar_io(conexion_inicial, cable_io)
            if conexion_final is not None:
                self.connect(cable_key, conexion_final, n1 = 2, n2 = 1)
                self.conectar_io(conexion_final, cable_io)

            tritubos_del_cable = tritubos[indice_cable]

            tritubo_name = vals_cable["@tramo"]
            if tritubo_name == 'SIN DATO' and "@subtramo" in vals_cable:
                tritubo_name = vals_cable["@subtramo"]
            elif tritubo_name == 'SIN DATO' and vals_cable["@oName"] != "SIN NOMBRE":
                tritubo_name = vals_cable["@oName"]
            tri_ocfg, tri_val, tri_n_id = self.get_ocfg(feature, "TRITUBO", tritubo_name, f_type="fme_line")
            tri_sidx = poblar_sidx(tri_val)
            contador_tritubos += len(tritubos_del_cable)
            for indice_tritubo in range(0, len(tritubos_del_cable)):

                coordenadas_tritubo = tritubos_del_cable[indice_tritubo][0]
                conexion_inicial_tri, conexion_final_tri = tritubos_del_cable[indice_tritubo][1]
                # escribimos
                tritubo_key = self.write_object(tri_n_id, "gc/duc", tri_val, tri_sidx, coordenadas_tritubo)
                if conexion_inicial_tri is not None:
                    self.connect(tritubo_key, conexion_inicial_tri, n1 = 1, n2 = 1)
                if conexion_final_tri is not None:
                    self.connect(tritubo_key, conexion_final_tri, n1 = 2, n2 = 1)
        print(f"Cantidad total de cables: {len(cables)}")
        print(f"Cantidad total de tritubos: {contador_tritubos}")

    def obtenerSoloCajas(self):
        cajas_id = []
        cajas_coord = []
        for i in range(0, len(self.cajas_id)):
            if self.tipo_camara[i] == "go/fo/em":
                cajas_id.append(self.cajas_id[i])
                cajas_coord.append(self.cajas_coord[i])
        return cajas_id, cajas_coord

    def fronterasPorProximidad(self):
        cajas_id, cajas_coord = self.obtenerSoloCajas()

        fronteras_coord = list(map(lambda x:(x.getCoordinate(0)[0], x.getCoordinate(0)[1]),
                                   self.fronteras))

        # hacemos los cKDTree
        tree_cajas = KDTree(cajas_coord)
        _, indexes = tree_cajas.query(fronteras_coord,k = 1)

        for i in range(0, len(indexes)):
            feature = self.fronteras[i]
            fo_name = feature.getAttribute("kml_name")
            if fo_name is None:
                continue
            fo_name = estandarizar(fo_name)
            key = cajas_id[indexes[i]]
            print(f"[INFO] Caja '{key}' FO de nombre {fo_name}")
            if key not in self.cajas_frontera:
                print(f"[INFO] Poniendo variable @frontera en true")
                self.ffval.write(f'SADD {key}:val "{self.logs}:@frontera|true|0"\n')
                self.cajas_frontera.append(key)
            self.ffval.write(f'SADD {key}:val "{self.logs}:@nombreFrontera|{fo_name}|0"\n')

    def obtener_tipo_nodo_edts(self, feature):
        acronimo_nodo = get_acronimo_nodo(feature)
        tipo = None
        if acronimo_nodo is not None:
            if acronimo_nodo in self.shelters:
                tipo = SHELTER[0]
            elif acronimo_nodo in self.gabinetes:
                tipo = GABINETE[0]
            elif acronimo_nodo in self.casetas:
                tipo = CASETA[0]
        return tipo

    def close(self):

        print("Cargando datos del EXCEL...")
        self.poblarDatosEntorno()
        self.poblarDatosRTU()
        self.poblarSitiosYNodos()
        self.poblarTramosYSubtramos()
        print(self.gabinetes)
        ff_name = "ARSAT"

        ffpath = RUTA_MIGRACIONES
        foldername = f"Migracion_{check_quotes(PROVINCIA.replace(' ','_'))}"

        path = ffpath + "\\" + foldername
        os.makedirs(path, exist_ok=True)

        print("Abriendo archivos de export...")

        self.ffo = open(path + "\O-" + ff_name + ".txt", "w")
        self.ffocfg = open(path + "\OCFG-" + ff_name + ".txt", "w")
        self.ffval = open(path + "\VAL-" + ff_name + ".txt", "w")
        self.ffsidx = open(path + "\SIDX-" + ff_name + ".txt", "w")
        self.ffv = open(path + "\V-" + ff_name + ".txt", "w")
        self.ffgeoidx = open(path + "\GEOIDX-" + ff_name + ".txt", "w")
        self.ffio = open(path + "\IO-" + ff_name + ".txt","w")
        self.ffco = open(path + "\CO-" + ff_name + ".txt", "w")

        print(f"Procesando KMZ, {len(self.featureList)} elementos restantes...")
        count = 0

        for feature in self.featureList:
            if count % 10 == 0:
                print(str(count)+"...")
            count+=1
            folder = self.get_root_folder(feature)
            if folder is None:
                continue
            # si es una frontera optica, la agregamos a la lista y pasamos al siguiente elemento
            if folder in SITIO or folder in RED:
                folder = self.get_folder(feature)

            f_id = feature.getAttribute("kml_id")
            f_id = check_quotes(f_id)
            f_name = feature.getAttribute("kml_name")
            if f_name is None:
                f_name = "Sin nombre"
            f_name = check_quotes(f_name)
            f_type = feature.getAttribute("fme_type")
            f_type = check_quotes(f_type)

            if folder in FRONTERA_OPTICA:
                self.fronteras.append(feature)
                continue

            ocfg = self.get_ocfg(feature,folder,f_name,f_type)

            coordinates = list(map(lambda x: (x[0], x[1]), feature.getAllCoordinates()))
            # si no tiene vertices, se saltea
            if coordinates is None:
                continue
            # verify that the coordinates are < 0
            if any(list(map(lambda x: x[0] > 0 or x[1] > 0,coordinates))):
                continue ##if a longitude is > 0, skip the element

            # antes que nada, si es un nodo escribimos el sitio
            if folder in GABINETE or folder in SHELTER or folder in CASETA:
                sitio_ocfg = self.get_ocfg(feature, SITIO[0], f_name, f_type)
                # si el sitio es None, es porque esta inactivo. Salteamos t o d o
                if sitio_ocfg is None:
                    continue
                sitio_val = sitio_ocfg[1]
                sitio_val["@kmlId"] = f_id
                sitio_sidx = poblar_sidx(sitio_val)

                self.write_object(sitio_ocfg[2], sitio_ocfg[0], sitio_val, sitio_sidx, coordinates)

            # si no es de ninguno de los tipos posibles, lo salteamos
            if ocfg is None:
                continue
            # si la linea no salio de la carpeta RED no la migro
            if ocfg[0] != "gc/fo" and f_type == "fme_line":
                continue
            val = ocfg[1]

            if val is None:
                val = {}
            if "@oName" not in val:
                val["@oName"] = "SIN NOMBRE"
            val["@kmlId"] = f_id
            sidx = poblar_sidx(val)

            if ocfg[0] == "go/fo/cam":
                val["@com"] = get_camara_comment(feature.getAttribute("kml_description"))

            key = self.write_object(ocfg[2], ocfg[0], val, sidx, coordinates)

            if ocfg[0] == "go/fo/em" or ocfg[0] == "go/fo/cam":
                self.cajas_coord.append(coordinates[0])
                self.cajas_id.append(key)
                self.tipo_camara.append(ocfg[0])
            if ocfg[0] == "go/fo/em":
                # varias cosas para hacer
                # primero, si es una frontera lo vamos a agregar a la lista de cajas frontera
                if "@frontera" in val:
                    self.cajas_frontera.append(key)
                # luego vamos a hacer la camara
                descripcion = feature.getAttribute("kml_description")
                if descripcion is None:
                    descripcion = "SIN NOMBRE"
                camara_name = get_camara_name(descripcion)
                camara_ocfg, camara_vals, camara_n_id  = \
                    self.get_ocfg(feature, CAMARA[0], name = camara_name, f_type = f_type)
                camara_vals["@tapas"] = "4"
                camara_vals["@tipoCam"] = "1"
                camara_vals['@aCaja'] = f_name
                camara_sidx = poblar_sidx(camara_vals)
                camara_key = self.write_object(camara_n_id, camara_ocfg, camara_vals, camara_sidx, coordinates)
                self.camara_por_caja[key] = camara_key

            if ocfg[0] == "go/fo/edts":
                sitio_ocfg = self.get_ocfg(feature, SITIO[0], name = f_name, f_type = f_type)
                new_coord = [move_coordinates(coordinates[0])]
                if sitio_ocfg is not None:
                    sitio_val = sitio_ocfg[1]
                    if sitio_val is None:
                        sitio_val = {}
                    if "@oName" not in sitio_val:
                        sitio_val["@oName"] = "SIN NOMBRE"
                    sitio_val["@kmlId"] = f_id
                    sitio_sidx = poblar_sidx(sitio_val)
                    self.write_object(sitio_ocfg[2], sitio_ocfg[0], sitio_val, sitio_sidx, new_coord)
                tipo = self.obtener_tipo_nodo_edts(feature)
                if tipo is not None:
                    nodo_ocfg = self.get_ocfg(feature, tipo, name = f_name, f_type = f_type)
                    if nodo_ocfg is not None:
                        nodo_val = nodo_ocfg[1]
                        if nodo_val is None:
                            nodo_val = {}
                        if "@oName" not in nodo_val:
                            nodo_val["@oName"] = "SIN NOMBRE"
                        nodo_val["@kmlId"] = f_id
                        nodo_sidx = poblar_sidx(nodo_val)
                        self.write_object(nodo_ocfg[2], nodo_ocfg[0], nodo_val, nodo_sidx, new_coord)


        print("Creando conexiones y dividiendo cables...")

        self.divide_cable()

        self.fronterasPorProximidad()

        print("Operacion completada.")

        self.actualizarIdMaximo(ffpath)

        close_files(self.ffo,
                    self.ffocfg,
                    self.ffval,
                    self.ffsidx,
                    self.ffv,
                    self.ffgeoidx,
                    self.ffio,
                    self.ffco)

        print("Generando loaddata...")
        generate_loaddata(path,
                          self.ffval,
                          self.ffo,
                          self.ffv,
                          self.ffocfg,
                          self.ffsidx,
                          self.ffgeoidx,
                          self.ffio,
                          self.ffco)
        print("Listo.")
        print(f"[INFO] Cantidad de cajas: {len(self.camara_por_caja)}")
        print(f"[INFO] Cantidad de camaras: {len(self.cajas_id)}")
        print(f"ID Maximo despues de la migracion: {self.object_id}")
        print(f"Sitios del excel que quedaron fuera: {list(self.sitios.keys())}")

class AgregadorDeVertices:
    # devuelve el indice del tritubo y el indice del vertice
    @staticmethod
    def __obtenerVerticeTritubo__(tritubo_cable, indice_vertice_cable):
        numero_de_vertice = 0
        for indice_tritubo in range(0, len(tritubo_cable)):
            vertices, camaras = tritubo_cable[indice_tritubo]
            # iteramos hasta el anteúltimo vertice
            # si matchearía con el último, va a matchear cuando encuentre el primero del siguiente tritubo
            for indice_vertice in range(0, len(vertices) - 1):
                if numero_de_vertice == indice_vertice_cable:
                    return indice_tritubo, indice_vertice
                numero_de_vertice += 1

    @staticmethod
    def __distancia_dentro_de_rango__(distancia):
        return distancia <= 0.0005

    @staticmethod
    def __get_distance__(coord_cable_inicio, coord_cable_fin, coord_caja):
        ls = LineString([(coord_cable_inicio[1], coord_cable_inicio[0]), (coord_cable_fin[1], coord_cable_fin[0])])
        p = Point((coord_caja[1], coord_caja[0]))
        pinicial = Point((coord_cable_inicio[1], coord_cable_inicio[0]))
        pfinal = Point((coord_cable_fin[1], coord_cable_fin[0]))
        # print(ls.distance(p))
        if pinicial.distance(p) < 0.00007 or pfinal.distance(p) < 0.00007:
            return None
        return ls.distance(p)

    @staticmethod
    def __corregir_cable_con_caja__(cables,
                                    tritubos,
                                    indice_cable,
                                    indice_vertice,
                                    caja_id,
                                    caja_coord,
                                    caja_camara):
        vertices = cables[indice_cable][0]
        caja_inicial_id, caja_final_id, feature, tipo_cable = cables[indice_cable][1]
        # ------------------ divido el cable -------------------

        vertices_iniciales = vertices[:indice_vertice + 1]
        vertices_iniciales.append(caja_coord)
        vertices_finales = vertices[indice_vertice + 1:]
        vertices_finales.insert(0, caja_coord)
        cajas_iniciales = caja_inicial_id, caja_id, feature, tipo_cable
        cajas_finales = caja_id, caja_final_id, feature, tipo_cable
        # print(f"CABLES BEFORE: {cables}")
        cables.pop(indice_cable)
        cables.insert(indice_cable, (vertices_iniciales, cajas_iniciales))
        cables.insert(indice_cable + 1, (vertices_finales, cajas_finales))
        # print(f"CABLES AFTER:  {cables}")
        # ---------------- divido el tritubo ---------------
        # print(f"TRITUBOS BEFORE: {tritubos}")
        tritubos_del_cable = tritubos[indice_cable]
        indice_tritubo, indice_vertice_tritubo = AgregadorDeVertices.__obtenerVerticeTritubo__(tritubos_del_cable,
                                                                                               indice_vertice)
        tritubo, camaras = tritubos_del_cable[indice_tritubo]
        tritubo_inicial = tritubo[:indice_vertice_tritubo + 1]
        tritubo_inicial.append(caja_coord)
        camaras_iniciales = camaras[0], caja_camara

        tritubo_final = tritubo[indice_vertice_tritubo + 1:]
        tritubo_final.insert(0, caja_coord)
        camaras_finales = caja_camara, camaras[1]

        tritubo_cable_anterior = tritubos_del_cable[: indice_tritubo + 1]
        tritubo_cable_anterior.pop(-1)
        tritubo_cable_anterior.append((tritubo_inicial, camaras_iniciales))
        tritubo_cable_siguiente = tritubos_del_cable[indice_tritubo + 1:]
        tritubo_cable_siguiente.insert(0, (tritubo_final, camaras_finales))

        tritubos.pop(indice_cable)
        tritubos.insert(indice_cable, tritubo_cable_anterior)
        tritubos.insert(indice_cable + 1, tritubo_cable_siguiente)
        # print(f"TRITUBOS AFTER:  {tritubos}")

    @staticmethod
    def __corregir_cable_con_camara__(cables,
                                      tritubos,
                                      indice_cable,
                                      indice_vertice,
                                      camara_id,
                                      camara_coord):
        vertices = cables[indice_cable][0]
        caja_inicial_id, caja_final_id, feature, tipo_cable = cables[indice_cable][1]
        # ------------------ agrego vertice al cable -------------------
        # print(f"CABLES BEFORE: {cables}")
        vertices.insert(indice_vertice + 1, camara_coord)
        cables.pop(indice_cable)
        cables.insert(indice_cable, (vertices, (caja_inicial_id, caja_final_id, feature, tipo_cable)))
        # print(f"CABLES AFTER:  {cables}")
        # ---------------- divido el tritubo ---------------
        tritubos_del_cable = tritubos[indice_cable]
        indices_tritubo = AgregadorDeVertices.__obtenerVerticeTritubo__(tritubos_del_cable,
                                                                        indice_vertice)
        # print(f"TRITUBOS BEFORE: {tritubos}")
        indice_tritubo, indice_vertice_tritubo = indices_tritubo[0], indices_tritubo[1]
        tritubo, camaras = tritubos_del_cable[indice_tritubo]
        tritubo_inicial = tritubo[:indice_vertice_tritubo + 1]
        tritubo_inicial.append(camara_coord)
        camaras_iniciales = camaras[0], camara_id
        tritubo_final = tritubo[indice_vertice_tritubo + 1:]
        tritubo_final.insert(0, camara_coord)
        camaras_finales = camara_id, camaras[1]
        tritubos[indice_cable].pop(indice_tritubo)
        tritubos[indice_cable].insert(indice_tritubo, (tritubo_inicial, camaras_iniciales))
        tritubos[indice_cable].insert(indice_tritubo + 1, (tritubo_final, camaras_finales))
        # print(f"TRITUBOS AFTER:  {tritubos}")

    @staticmethod
    def corregir_cables(cajas_coord, cajas_id, cajas_tipo, camara_por_caja, cables, tritubos):
        contador_matches = 0
        for indice_caja in range(0, len(cajas_coord)):

            caja_id = cajas_id[indice_caja]
            caja_coord = cajas_coord[indice_caja]
            caja_tipo = cajas_tipo[indice_caja]

            print(f"PROCESANDO CAJA {cajas_id[indice_caja]}")

            indice_cable_mas_cercano = None
            indice_vertice_minima = None
            distancia_minima = None

            # por cada cable, procesamos la distancia a cada par de vertices y guardamos los datos de la minima
            for indice_cable in range(0, len(cables)):
                vertices = cables[indice_cable][0]

                # por cada par de vertices
                for indice_vertice in range(0, len(vertices) - 1):
                    # el metodo puede devolver None si se encuentra muy cercano a los vertices
                    distancia = AgregadorDeVertices.__get_distance__(vertices[indice_vertice],
                                                                     vertices[indice_vertice + 1],
                                                                     caja_coord)
                    if distancia is not None and (distancia_minima is None or distancia < distancia_minima):
                        distancia_minima = distancia
                        indice_cable_mas_cercano = indice_cable
                        indice_vertice_minima = indice_vertice

            # una vez que procesamos todos los pares de vertices de todos los cables, si el mas cercano pasa
            # el chequeo de distancias, corregimos el cable

            if AgregadorDeVertices.__distancia_dentro_de_rango__(distancia_minima):

                print(f"INDICE {indice_cable_mas_cercano} CON LA CAJA {cajas_id[indice_caja]}")

                contador_matches += 1
                # si es una caja hay que dividir el cable y el tritubo
                if caja_tipo == "go/fo/em":
                    caja_camara = camara_por_caja[caja_id]
                    AgregadorDeVertices.__corregir_cable_con_caja__(cables,
                                                                    tritubos,
                                                                    indice_cable_mas_cercano,
                                                                    indice_vertice_minima,
                                                                    caja_id,
                                                                    caja_coord,
                                                                    caja_camara)
                # si no es una caja, es una camara. Hay que dividir el tritubo
                else:
                    AgregadorDeVertices.__corregir_cable_con_camara__(cables,
                                                                    tritubos,
                                                                    indice_cable_mas_cercano,
                                                                    indice_vertice_minima,
                                                                    caja_id,
                                                                    caja_coord)
        print(f"MATCHES: {contador_matches}")
